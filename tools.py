import base64
import email
import re

from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build

import storage
from features.web_search import search_web as _search_web
from features.weather import get_weather as _get_weather
from features.news import get_news as _get_news
from features.bland_ai import make_phone_call as _make_call, get_call_status as _call_status, list_recent_calls as _recent_calls
from features.reservations import search_opentable, search_thefork, find_business_phone


def get_services(creds: Credentials):
    gmail = build("gmail", "v1", credentials=creds)
    calendar = build("calendar", "v3", credentials=creds)
    people = build("people", "v1", credentials=creds)
    drive = build("drive", "v3", credentials=creds)
    return gmail, calendar, people, drive


# ── EMAIL ────────────────────────────────────────────────────────────────────

def get_new_message_ids(gmail) -> list[str]:
    """Utilise Gmail History API pour récupérer les IDs des nouveaux messages
    depuis le dernier check. Beaucoup plus rapide que de re-scanner."""
    last_hid = storage.get_last_history_id()
    if not last_hid:
        # Premier appel : on initialise le historyId sans retourner de messages
        profile = gmail.users().getProfile(userId="me").execute()
        storage.set_last_history_id(profile["historyId"])
        return []
    try:
        response = gmail.users().history().list(
            userId="me", startHistoryId=last_hid,
            historyTypes=["messageAdded"],
        ).execute()
        storage.set_last_history_id(response.get("historyId", last_hid))
        new_ids = set()
        for record in response.get("history", []):
            for added in record.get("messagesAdded", []):
                msg = added["message"]
                # Ignorer les brouillons et les messages envoyés
                labels = msg.get("labelIds", [])
                if "INBOX" in labels:
                    new_ids.add(msg["id"])
        return list(new_ids)
    except Exception:
        # historyId expiré ou erreur → reset
        profile = gmail.users().getProfile(userId="me").execute()
        storage.set_last_history_id(profile["historyId"])
        return []


def list_emails(gmail, max_results=10, query="is:unread"):
    result = gmail.users().messages().list(
        userId="me", maxResults=max_results, q=query
    ).execute()
    messages = result.get("messages", [])
    summaries = []
    for msg in messages:
        m = gmail.users().messages().get(
            userId="me", id=msg["id"], format="metadata",
            metadataHeaders=["Subject", "From", "Date"]
        ).execute()
        headers = {h["name"]: h["value"] for h in m["payload"]["headers"]}
        summaries.append({
            "id": msg["id"],
            "subject": headers.get("Subject", "(sans objet)"),
            "from": headers.get("From", ""),
            "date": headers.get("Date", ""),
            "snippet": m.get("snippet", ""),
        })
    return summaries


def _extract_body(payload: dict) -> str:
    def _strip_html(html: str) -> str:
        text = re.sub(r"<[^>]+>", " ", html)
        return re.sub(r"\s{2,}", " ", text).strip()

    def _decode_part(part: dict) -> str:
        data = part.get("body", {}).get("data", "")
        return base64.urlsafe_b64decode(data).decode("utf-8", errors="ignore") if data else ""

    parts = payload.get("parts", [])
    if not parts:
        parts = [payload]

    plain, html = "", ""
    for part in parts:
        mime = part.get("mimeType", "")
        if mime == "text/plain":
            plain += _decode_part(part)
        elif mime == "text/html":
            html += _decode_part(part)
        elif mime.startswith("multipart/"):
            sub = _extract_body(part)
            if sub:
                plain += sub

    return (plain or _strip_html(html))[:4000]


def _list_attachments(payload: dict) -> list[dict]:
    """Extrait la liste des pièces jointes d'un email."""
    attachments = []
    parts = payload.get("parts", [])
    for part in parts:
        filename = part.get("filename", "")
        if filename and part.get("body", {}).get("attachmentId"):
            attachments.append({
                "id": part["body"]["attachmentId"],
                "filename": filename,
                "mimeType": part.get("mimeType", ""),
                "size": part.get("body", {}).get("size", 0),
            })
        if part.get("parts"):
            attachments.extend(_list_attachments(part))
    return attachments


def read_email(gmail, email_id: str):
    msg = gmail.users().messages().get(
        userId="me", id=email_id, format="full"
    ).execute()
    headers = {h["name"]: h["value"] for h in msg["payload"]["headers"]}
    body = _extract_body(msg["payload"])
    attachments = _list_attachments(msg["payload"])
    return {
        "id": email_id,
        "thread_id": msg.get("threadId"),
        "subject": headers.get("Subject"),
        "from": headers.get("From"),
        "to": headers.get("To"),
        "date": headers.get("Date"),
        "body": body,
        "attachments": attachments,
    }


def get_attachment(gmail, email_id: str, attachment_id: str) -> str:
    """Télécharge une pièce jointe et retourne son contenu texte (si lisible)."""
    att = gmail.users().messages().attachments().get(
        userId="me", messageId=email_id, id=attachment_id
    ).execute()
    data = base64.urlsafe_b64decode(att["data"])
    # Tenter de décoder en texte
    try:
        text = data.decode("utf-8", errors="ignore")
        return text[:4000]
    except Exception:
        return f"[Pièce jointe binaire, {len(data)} octets — contenu non lisible en texte]"


def send_email(gmail, to: str, subject: str, body: str):
    message = email.message.EmailMessage()
    message["To"] = to
    message["Subject"] = subject
    message.set_content(body)
    raw = base64.urlsafe_b64encode(message.as_bytes()).decode()
    gmail.users().messages().send(userId="me", body={"raw": raw}).execute()
    return "Email envoyé avec succès."


def reply_email(gmail, email_id: str, body: str):
    original = gmail.users().messages().get(
        userId="me", id=email_id, format="metadata",
        metadataHeaders=["Subject", "From", "Message-ID"]
    ).execute()
    headers = {h["name"]: h["value"] for h in original["payload"]["headers"]}
    msg = email.message.EmailMessage()
    msg["To"] = headers.get("From")
    msg["Subject"] = "Re: " + headers.get("Subject", "")
    msg["In-Reply-To"] = headers.get("Message-ID", "")
    msg.set_content(body)
    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    gmail.users().messages().send(
        userId="me",
        body={"raw": raw, "threadId": original["threadId"]}
    ).execute()
    return "Réponse envoyée."


def create_draft(gmail, to: str, subject: str, body: str):
    message = email.message.EmailMessage()
    message["To"] = to
    message["Subject"] = subject
    message.set_content(body)
    raw = base64.urlsafe_b64encode(message.as_bytes()).decode()
    draft = gmail.users().drafts().create(
        userId="me", body={"message": {"raw": raw}}
    ).execute()
    return f"Brouillon créé (ID: {draft['id']})."


def archive_email(gmail, email_id: str):
    gmail.users().messages().modify(
        userId="me", id=email_id, body={"removeLabelIds": ["INBOX"]}
    ).execute()
    return "Email archivé."


def trash_email(gmail, email_id: str):
    gmail.users().messages().trash(userId="me", id=email_id).execute()
    return "Email mis à la corbeille."


def bulk_trash(gmail, query: str) -> str:
    """Met à la corbeille tous les emails correspondant à la query, par batch."""
    total = 0
    page_token = None
    while True:
        result = gmail.users().messages().list(
            userId="me", q=query, maxResults=500, pageToken=page_token
        ).execute()
        ids = [m["id"] for m in result.get("messages", [])]
        if not ids:
            break
        for email_id in ids:
            gmail.users().messages().trash(userId="me", id=email_id).execute()
        total += len(ids)
        page_token = result.get("nextPageToken")
        if not page_token:
            break
    return f"{total} emails mis à la corbeille."


def mark_all_read(gmail, query: str = "is:unread") -> str:
    """Marque tous les emails correspondant à la query comme lus, par batch."""
    total = 0
    page_token = None
    while True:
        result = gmail.users().messages().list(
            userId="me", q=query, maxResults=500, pageToken=page_token
        ).execute()
        ids = [m["id"] for m in result.get("messages", [])]
        if not ids:
            break
        gmail.users().messages().batchModify(
            userId="me",
            body={"ids": ids, "removeLabelIds": ["UNREAD"]}
        ).execute()
        total += len(ids)
        page_token = result.get("nextPageToken")
        if not page_token:
            break
    return f"{total} emails marqués comme lus."


def mark_email(gmail, email_id: str, read: bool = True):
    body = {"removeLabelIds": ["UNREAD"]} if read else {"addLabelIds": ["UNREAD"]}
    gmail.users().messages().modify(userId="me", id=email_id, body=body).execute()
    return f"Email marqué comme {'lu' if read else 'non lu'}."


def set_vacation_reply(gmail, message: str, subject: str = "Absent(e) du bureau",
                       start_date: str = None, end_date: str = None):
    body = {
        "enableAutoReply": True,
        "responseSubject": subject,
        "responseBodyPlainText": message,
        "restrictToContacts": False,
        "restrictToDomain": False,
    }
    if start_date:
        from datetime import datetime
        dt = datetime.fromisoformat(start_date)
        body["startTime"] = str(int(dt.timestamp() * 1000))
    if end_date:
        from datetime import datetime
        dt = datetime.fromisoformat(end_date)
        body["endTime"] = str(int(dt.timestamp() * 1000))
    gmail.users().settings().updateVacation(userId="me", body=body).execute()
    return "Réponse automatique activée."


def disable_vacation_reply(gmail):
    gmail.users().settings().updateVacation(
        userId="me", body={"enableAutoReply": False}
    ).execute()
    return "Réponse automatique désactivée."


# ── CALENDRIER ───────────────────────────────────────────────────────────────

def list_calendars(calendar) -> list:
    """Liste tous les calendriers accessibles."""
    result = calendar.calendarList().list().execute()
    return [
        {
            "id": c["id"],
            "name": c.get("summary", ""),
            "primary": c.get("primary", False),
            "color": c.get("backgroundColor", ""),
        }
        for c in result.get("items", [])
    ]


def list_events(calendar, max_results=10, time_min=None, calendar_id="primary"):
    from datetime import datetime, timezone
    if not time_min:
        time_min = datetime.now(timezone.utc).isoformat()
    result = calendar.events().list(
        calendarId=calendar_id, timeMin=time_min,
        maxResults=max_results, singleEvents=True, orderBy="startTime"
    ).execute()
    events = []
    for e in result.get("items", []):
        start = e["start"].get("dateTime", e["start"].get("date"))
        end = e["end"].get("dateTime", e["end"].get("date"))
        events.append({
            "id": e["id"],
            "title": e.get("summary", "(sans titre)"),
            "start": start,
            "end": end,
            "location": e.get("location", ""),
            "description": e.get("description", ""),
            "attendees": [a.get("email") for a in e.get("attendees", [])],
            "calendar": calendar_id,
        })
    return events


def create_event(calendar, title: str, start: str, end: str,
                 description: str = "", location: str = "", calendar_id: str = "primary"):
    event = {
        "summary": title,
        "description": description,
        "location": location,
        "start": {"dateTime": start, "timeZone": "Europe/Paris"},
        "end": {"dateTime": end, "timeZone": "Europe/Paris"},
    }
    result = calendar.events().insert(calendarId=calendar_id, body=event).execute()
    return f"Événement créé : {result.get('htmlLink')}"


def update_event(calendar, event_id: str, title: str = None, start: str = None,
                 end: str = None, description: str = None, location: str = None,
                 calendar_id: str = "primary"):
    event = calendar.events().get(calendarId=calendar_id, eventId=event_id).execute()
    if title:
        event["summary"] = title
    if description is not None:
        event["description"] = description
    if location is not None:
        event["location"] = location
    if start:
        event["start"] = {"dateTime": start, "timeZone": "Europe/Paris"}
    if end:
        event["end"] = {"dateTime": end, "timeZone": "Europe/Paris"}
    result = calendar.events().update(
        calendarId=calendar_id, eventId=event_id, body=event
    ).execute()
    return f"Événement mis à jour : {result.get('summary')}"


def delete_event(calendar, event_id: str, calendar_id: str = "primary"):
    calendar.events().delete(calendarId=calendar_id, eventId=event_id).execute()
    return "Événement supprimé."


# ── CONTACTS ─────────────────────────────────────────────────────────────────

def search_contacts(people, query: str) -> list:
    result = people.people().searchContacts(
        query=query,
        readMask="names,emailAddresses,phoneNumbers,organizations"
    ).execute()
    contacts = []
    for item in result.get("results", []):
        p = item.get("person", {})
        name = p.get("names", [{}])[0].get("displayName", "")
        emails = [e["value"] for e in p.get("emailAddresses", [])]
        phones = [ph["value"] for ph in p.get("phoneNumbers", [])]
        org = p.get("organizations", [{}])[0].get("name", "")
        contacts.append({"nom": name, "emails": emails, "telephones": phones, "organisation": org})
    return contacts


# ── GOOGLE DRIVE ─────────────────────────────────────────────────────────────

def search_drive(drive, query: str, max_results: int = 5) -> list:
    safe_query = query.replace("\\", "\\\\").replace("'", "\\'")
    result = drive.files().list(
        q=f"fullText contains '{safe_query}' and trashed=false",
        pageSize=max_results,
        fields="files(id, name, mimeType, modifiedTime, webViewLink)"
    ).execute()
    return result.get("files", [])


def read_drive_file(drive, file_id: str) -> str:
    file = drive.files().get(fileId=file_id, fields="mimeType,name").execute()
    mime = file.get("mimeType", "")
    name = file.get("name", "")
    if "google-apps.document" in mime:
        content = drive.files().export(fileId=file_id, mimeType="text/plain").execute()
        return content.decode("utf-8")[:4000]
    elif mime.startswith("text/"):
        content = drive.files().get_media(fileId=file_id).execute()
        return content.decode("utf-8")[:4000]
    else:
        return f"Fichier '{name}' (type : {mime}) — non lisible directement. Ouvre-le sur Drive."


def create_drive_doc(drive, title: str, content: str = "", folder_id: str = None) -> str:
    """Crée un Google Doc avec du contenu texte."""
    metadata = {
        "name": title,
        "mimeType": "application/vnd.google-apps.document",
    }
    if folder_id:
        metadata["parents"] = [folder_id]
    from googleapiclient.http import MediaInMemoryUpload
    media = MediaInMemoryUpload(
        content.encode("utf-8"), mimetype="text/plain", resumable=False
    )
    file = drive.files().create(
        body=metadata, media_body=media, fields="id,webViewLink"
    ).execute()
    return f"Document créé : {file.get('webViewLink', file['id'])}"


def update_drive_doc(drive, file_id: str, content: str) -> str:
    """Remplace le contenu d'un Google Doc existant."""
    from googleapiclient.http import MediaInMemoryUpload
    media = MediaInMemoryUpload(
        content.encode("utf-8"), mimetype="text/plain", resumable=False
    )
    drive.files().update(fileId=file_id, media_body=media).execute()
    return f"Document mis à jour."


# ── RECHERCHE / MÉTÉO / ACTUALITÉS ───────────────────────────────────────────

def search_web(query: str, max_results: int = 5) -> list:
    return _search_web(query, max_results)


def get_weather(city: str = "Paris") -> dict:
    return _get_weather(city)


def get_news(topic: str = "france", max_results: int = 5) -> list:
    return _get_news(topic, max_results)


# ── EXPORT DÉPENSES → EXCEL SUR DRIVE ───────────────────────────────────────

def export_expenses_to_drive(drive, month: str = None) -> str:
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from googleapiclient.http import MediaIoBaseUpload

    expenses = storage.list_expenses(month)
    if not expenses:
        return "Aucune dépense à exporter."

    wb = openpyxl.Workbook()
    ws = wb.active
    label = month if month else "Toutes"
    ws.title = f"Dépenses {label}"

    # ── En-têtes ──
    headers = ["#", "Date", "Montant (€)", "Catégorie", "Description"]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    ws.row_dimensions[1].height = 22

    # ── Données ──
    alt_fill = PatternFill("solid", fgColor="EBF3FB")
    total = 0
    for row_idx, e in enumerate(expenses, 2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        values = [e["id"], e["date"], e["amount"], e["category"], e.get("description", "")]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = fill
            cell.border = border
            if col == 3:
                cell.number_format = '#,##0.00 "€"'
                cell.alignment = Alignment(horizontal="right")
        total += e["amount"]

    # ── Ligne total ──
    total_row = len(expenses) + 2
    ws.cell(row=total_row, column=2, value="TOTAL").font = Font(bold=True)
    total_cell = ws.cell(row=total_row, column=3, value=total)
    total_cell.font = Font(bold=True, color="1F4E79")
    total_cell.number_format = '#,##0.00 "€"'
    total_cell.alignment = Alignment(horizontal="right")

    # ── Largeurs colonnes ──
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 35

    # ── Upload sur Drive ──
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"Dépenses_{label}.xlsx"
    file_meta = {
        "name": filename,
        "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    media = MediaIoBaseUpload(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=False,
    )
    result = drive.files().create(
        body=file_meta, media_body=media, fields="id,webViewLink"
    ).execute()

    return f"Fichier Excel créé sur Drive : {result.get('webViewLink')}"


# ── NOTES ────────────────────────────────────────────────────────────────────

def add_note(content: str) -> str:
    return storage.add_note(content)

def list_notes() -> list:
    return storage.list_notes()

def delete_note(note_id: int) -> str:
    return storage.delete_note(note_id)


# ── TÂCHES ───────────────────────────────────────────────────────────────────

def add_task(title: str, due_date: str = None) -> str:
    return storage.add_task(title, due_date)

def list_tasks(include_done: bool = False) -> list:
    return storage.list_tasks(include_done)

def complete_task(task_id: int) -> str:
    return storage.complete_task(task_id)

def delete_task(task_id: int) -> str:
    return storage.delete_task(task_id)


# ── DÉPENSES ─────────────────────────────────────────────────────────────────

def add_expense(amount: float, category: str, description: str = "") -> str:
    return storage.add_expense(amount, category, description)

def list_expenses(month: str = None) -> list:
    return storage.list_expenses(month)

def expense_summary(month: str = None) -> dict:
    return storage.expense_summary(month)

def delete_expense(expense_id: int) -> str:
    return storage.delete_expense(expense_id)


# ── TEMPLATES EMAIL ──────────────────────────────────────────────────────────

def save_email_template(name: str, subject: str, body: str) -> str:
    return storage.save_email_template(name, subject, body)

def list_email_templates() -> list:
    return storage.list_email_templates()

def delete_email_template(template_id: int) -> str:
    return storage.delete_email_template(template_id)

def send_from_template(gmail, template_id: int, to: str, extra: str = "") -> str:
    tpl = storage.get_email_template(template_id)
    if not tpl:
        return f"Template #{template_id} introuvable."
    body = tpl["body"]
    if extra:
        body += f"\n\n{extra}"
    return send_email(gmail, to, tpl["subject"], body)


# ── ALERTES EMAIL ────────────────────────────────────────────────────────────

def add_email_alert(sender: str = "", keyword: str = "") -> str:
    return storage.add_email_alert(sender, keyword)

def list_email_alerts() -> list:
    return storage.list_email_alerts()

def remove_email_alert(alert_id: int) -> str:
    return storage.remove_email_alert(alert_id)


# ── APPELS TÉLÉPHONIQUES (BLAND.AI) ──────────────────────────────────────────

def make_phone_call(phone_number: str, task: str, max_duration: int = 5) -> dict:
    return _make_call(phone_number, task, max_duration)

def get_call_status(call_id: str) -> dict:
    return _call_status(call_id)

def list_recent_calls(limit: int = 5) -> list:
    return _recent_calls(limit)


# ── RÉSERVATIONS EN LIGNE ────────────────────────────────────────────────────

def reserve_opentable(restaurant: str, city: str, date: str,
                      time: str, guests: int = 2) -> dict:
    return search_opentable(restaurant, city, date, time, guests)

def reserve_thefork(restaurant: str, city: str, date: str,
                    time: str, guests: int = 2) -> dict:
    return search_thefork(restaurant, city, date, time, guests)

def find_phone_number(business_name: str, city: str) -> dict:
    return find_business_phone(business_name, city)


# ── MÉMOIRE ──────────────────────────────────────────────────────────────────

def save_memory(key: str, value: str) -> str:
    return storage.save_memory(key, value)

def get_memory() -> dict:
    return storage.get_memory()

def delete_memory(key: str) -> str:
    return storage.delete_memory(key)


# ── AUTOMATIONS ──────────────────────────────────────────────────────────────

def add_automation(user_id: int, description: str, schedule: str,
                   instruction: str, condition: str = None) -> str:
    automation = storage.add_automation(user_id, description, schedule, instruction, condition)
    msg = f"Automation #{automation['id']} créée : {description} ({schedule})"
    if condition:
        msg += f" [condition: {condition}]"
    return msg + "."

def list_automations(user_id: int) -> list:
    return storage.list_automations(user_id)

def delete_automation(auto_id: int) -> str:
    return storage.delete_automation(auto_id)


# ── SUIVI DE RELANCE EMAIL ───────────────────────────────────────────────────

def track_followup(user_id: int, email_id: str, thread_id: str, to: str,
                   subject: str, followup_days: int = 3) -> str:
    followup = storage.add_followup(user_id, email_id, thread_id, to, subject, followup_days)
    return f"Suivi #{followup['id']} activé : relance dans {followup_days} jours si pas de réponse de {to}."

def list_followups(user_id: int) -> list:
    return storage.list_followups(user_id)

def cancel_followup(followup_id: int) -> str:
    return storage.cancel_followup(followup_id)


# ── RAPPELS ──────────────────────────────────────────────────────────────────

def set_reminder(user_id: int, message: str, remind_at: str) -> str:
    reminder = storage.add_reminder(user_id, message, remind_at)
    return f"Rappel #{reminder['id']} programmé pour le {remind_at}."

def list_reminders(user_id: int) -> list:
    return storage.list_reminders(user_id)

def cancel_reminder(reminder_id: int) -> str:
    return storage.cancel_reminder(reminder_id)


# ── SURVEILLANCE WEB ────────────────────────────────────────────────────────

def add_web_monitor(user_id: int, url: str, description: str,
                    css_selector: str = None, interval_minutes: int = 60) -> str:
    monitor = storage.add_web_monitor(user_id, url, description, css_selector, interval_minutes)
    return f"Surveillance #{monitor['id']} créée : {description} (toutes les {interval_minutes} min)."

def list_web_monitors(user_id: int) -> list:
    return storage.list_web_monitors(user_id)

def delete_web_monitor(monitor_id: int) -> str:
    return storage.delete_web_monitor(monitor_id)

def check_web_page(url: str, css_selector: str = None) -> dict:
    from features.web_monitor import fetch_page_text
    return fetch_page_text(url, css_selector)


# ── RÉSUMÉ DE LIENS ─────────────────────────────────────────────────────────

def summarize_url(url: str) -> dict:
    from features.url_summarizer import extract_article_text
    return extract_article_text(url)


# ── SMS / WHATSAPP ──────────────────────────────────────────────────────────

def send_sms(to: str, body: str) -> dict:
    from features.twilio_sms import send_sms as _send
    return _send(to, body)

def send_whatsapp(to: str, body: str) -> dict:
    from features.twilio_sms import send_whatsapp as _send
    return _send(to, body)


# ── GITHUB ──────────────────────────────────────────────────────────────────

def github_repo_status(owner: str, repo: str) -> dict:
    from features.github_monitor import get_repo_status
    return get_repo_status(owner, repo)

def github_notifications() -> list:
    from features.github_monitor import list_notifications
    return list_notifications()

def add_github_repo(user_id: int, owner: str, repo: str) -> str:
    return storage.add_github_repo(user_id, owner, repo)

def list_github_repos(user_id: int) -> list:
    return storage.list_github_repos(user_id)

def delete_github_repo(repo_id: int) -> str:
    return storage.delete_github_repo(repo_id)


# ── BUDGET ──────────────────────────────────────────────────────────────────

def set_budget(category: str, amount: float, month: str = None) -> str:
    return storage.set_budget(category, amount, month)

def get_budgets() -> dict:
    return storage.get_budgets()

def delete_budget(category: str, month: str = None) -> str:
    return storage.delete_budget(category, month)

def check_budget_alerts(month: str = None) -> list:
    return storage.check_budget_alerts(month)


# ── MAPS & TRAJETS (OSM/OSRM, 100% gratuit) ────────────────────────────────

def geocode_address(address: str) -> dict:
    from features.maps import geocode
    return geocode(address)

def travel_time(origin: str, destination: str, mode: str = "driving") -> dict:
    from features.maps import travel_time as _tt
    return _tt(origin, destination, mode)

def get_directions(origin: str, destination: str, mode: str = "driving") -> dict:
    from features.maps import get_directions as _gd
    return _gd(origin, destination, mode)

def when_to_leave(destination: str, arrival_time_iso: str, origin: str = None,
                  mode: str = "driving", buffer_minutes: int = 10) -> dict:
    from features.maps import should_leave_by
    return should_leave_by(destination, arrival_time_iso, origin, mode, buffer_minutes)


# ── RAG MÉMOIRE SÉMANTIQUE (ChromaDB local, 100% gratuit) ──────────────────

def search_memory(query: str, top_k: int = 5, source_filter: str = None) -> list:
    from features.rag import search_memory as _search
    return _search(query, top_k, source_filter)

def rag_stats() -> dict:
    from features.rag import get_stats
    return get_stats()

def index_recent_content(user_id: int) -> dict:
    """Indexe les dernières conversations, notes et mémoire de l'utilisateur."""
    from features.rag import index_documents
    docs = []
    # Conversations
    conv = storage.load_conversation(user_id)
    for m in conv:
        if isinstance(m.get("content"), str):
            docs.append({
                "content": m["content"],
                "source": f"conversation_{m['role']}",
                "timestamp": "",
            })
    # Notes
    for n in storage.list_notes():
        docs.append({
            "content": n["content"],
            "source": "note",
            "timestamp": n.get("created_at", ""),
        })
    # Mémoire clé-valeur
    memory = storage.get_memory()
    for k, v in memory.items():
        docs.append({
            "content": f"{k} : {v}",
            "source": "memory_kv",
            "timestamp": "",
        })
    count = index_documents(docs)
    return {"indexed": count, "total_docs": len(docs)}
